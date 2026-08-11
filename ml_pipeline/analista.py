#!/usr/bin/env python3
"""
Analista de anúncios do Mercado Livre.

Lê os dados coletados pelo n8n (dados_conta.json, com TODAS as imagens de cada
anúncio), cruza com as métricas REAIS exportadas do ML e com as regras da Central
de Aprendizagem, e gera planos de ação por anúncio e para a conta — incluindo
cupons, promoções, kits e estrutura de anúncio.

Chama a API do Claude com visão (imagens em base64 — a única forma que funciona
com o CDN mlstatic.com). NÃO recalcula conversão: usa a métrica real fornecida.

Uso:
    export ANTHROPIC_API_KEY=...            # ou: ant auth login
    python analista.py                      # usa os caminhos padrão abaixo
    python analista.py --limite 3           # testa com os 3 primeiros anúncios
"""

from __future__ import annotations

import argparse
import base64
import os
import json
import sys
import time
from pathlib import Path

import anthropic
import requests

# ----------------------------------------------------------------------------
# Config — ajuste aqui
# ----------------------------------------------------------------------------
MODELO = "claude-sonnet-5"        # troque para "claude-opus-4-8" (mais capaz) ou "claude-haiku-4-5" (mais barato)
EFFORT = "high"                    # low | medium | high | xhigh | max
MAX_IMGS_POR_ANUNCIO = 5           # capa + 4 infográficos; sobe/desce conforme custo
MAX_TOKENS = 12000                 # < 16k → sem streaming; margem p/ o tokenizador do Sonnet 5

BASE = Path(__file__).resolve().parent
PADRAO_DADOS = BASE / "dados_conta.json"
PADRAO_REGRAS = BASE / "regras_central.md"
# Relatório do ML com as métricas de verdade (opcional, mas recomendado):
PADRAO_METRICAS = Path.home() / "Downloads" / "relatorios ml" / "Desempenho dos seus anúncios.xlsx"
SAIDA_DIR = BASE / "saida"

# ----------------------------------------------------------------------------
# Schemas de saída (JSON Schema — força estrutura consistente entre todos os itens)
# ----------------------------------------------------------------------------
SCHEMA_ANUNCIO = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "item_id": {"type": "string"},
        "titulo_atual": {"type": "string"},
        "notas_pilares": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "imagens": {"type": "integer"},
                "titulo": {"type": "integer"},
                "ficha_tecnica": {"type": "integer"},
                "descricao": {"type": "integer"},
            },
            "required": ["imagens", "titulo", "ficha_tecnica", "descricao"],
        },
        "nota_geral": {"type": "number"},
        "gargalo": {"type": "string", "enum": ["atracao", "conversao", "ambos", "indefinido"]},
        "diagnostico": {"type": "string"},
        "problemas": {"type": "array", "items": {"type": "string"}},
        "duvidas_recorrentes": {"type": "array", "items": {"type": "string"}},
        "titulo_sugerido": {"type": "string"},
        "acoes": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "ordem": {"type": "integer"},
                    "acao": {"type": "string"},
                    "tipo": {"type": "string", "enum": [
                        "titulo", "fotos", "ficha", "descricao", "preco", "cupom",
                        "promocao", "kit", "catalogo", "variacao", "full", "ads",
                        "clip", "estoque", "outro",
                    ]},
                    "impacto_esperado": {"type": "string"},
                    "esforco": {"type": "string"},
                },
                "required": ["ordem", "acao", "tipo", "impacto_esperado", "esforco"],
            },
        },
        "prioridade": {"type": "string", "enum": ["alta", "media", "baixa"]},
    },
    "required": [
        "item_id", "titulo_atual", "notas_pilares", "nota_geral", "gargalo",
        "diagnostico", "problemas", "duvidas_recorrentes", "titulo_sugerido",
        "acoes", "prioridade",
    ],
}

SCHEMA_CONTA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "resumo": {"type": "string"},
        "jogadas": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "tipo": {"type": "string", "enum": [
                        "cupom", "promocao", "kit", "full", "ads", "minha_pagina",
                        "preco", "catalogo", "reputacao", "recompra", "fiscal", "outro",
                    ]},
                    "titulo": {"type": "string"},
                    "descricao": {"type": "string"},
                    "itens_alvo": {"type": "array", "items": {"type": "string"}},
                    "impacto_esperado": {"type": "string"},
                    "prioridade": {"type": "string", "enum": ["alta", "media", "baixa"]},
                },
                "required": ["tipo", "titulo", "descricao", "itens_alvo", "impacto_esperado", "prioridade"],
            },
        },
        "itens_prioritarios": {"type": "array", "items": {"type": "string"}},
    },
    "required": ["resumo", "jogadas", "itens_prioritarios"],
}

# Identificação da conta e do catálogo. Ficam fora do código para o script
# servir a qualquer vendedor — e para não versionar dado de uma conta real.
CONTA = os.getenv("ML_CONTA", "a conta")
PERFIL_CATALOGO = os.getenv(
    "ML_PERFIL_CATALOGO",
    "descreva aqui as categorias que a conta vende",
)

SYSTEM_PROMPT_BASE = f"""Você é um consultor sênior de Mercado Livre analisando {CONTA} \
({PERFIL_CATALOGO}).

Seu trabalho: para cada anúncio, olhar TODAS as fotos fornecidas, o título, a ficha técnica, a \
descrição e as perguntas dos compradores, cruzar com as métricas REAIS do painel do ML (fornecidas \
no texto) e com as regras da Central abaixo, e produzir um diagnóstico honesto + plano de ação.

Regras de ouro:
- NÃO calcule conversão. Use exatamente o número real fornecido. Se não houver, escreva o gargalo como "indefinido".
- Classifique o gargalo pelo funil real: muitas visitas + baixa conversão = "conversao"; poucas visitas + \
boa conversão = "atracao". As ações mudam conforme o gargalo.
- Baseie a nota de imagens SOMENTE nas fotos fornecidas (você vê a galeria real, não só a capa).
- Se as perguntas revelarem dúvidas recorrentes (ex.: espessura, roldana, compatibilidade), liste-as e \
mande levar a resposta para a ficha/descrição.
- Seja concreto e realista nos impactos; nada de porcentagens inventadas com falsa precisão.

--- REGRAS DA CENTRAL DE APRENDIZAGEM ---
{regras}
"""


# ----------------------------------------------------------------------------
# Utilidades
# ----------------------------------------------------------------------------
def detectar_media_type(data: bytes) -> str | None:
    """Descobre o tipo da imagem pelos magic bytes (mlstatic mistura webp/jpg)."""
    if data[:3] == b"\xff\xd8\xff":
        return "image/jpeg"
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return "image/png"
    if data[:4] == b"GIF8":
        return "image/gif"
    if data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        return "image/webp"
    return None


def baixar_imagens(urls: list[str], limite: int) -> list[dict]:
    """Baixa e converte imagens em blocos de imagem base64 para a API."""
    blocos = []
    for url in urls[:limite]:
        try:
            r = requests.get(url, timeout=30)
            r.raise_for_status()
            mt = detectar_media_type(r.content)
            if not mt:
                print(f"    (imagem ignorada, tipo desconhecido: {url})", file=sys.stderr)
                continue
            blocos.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": mt,
                    "data": base64.standard_b64encode(r.content).decode("ascii"),
                },
            })
        except Exception as e:
            print(f"    (falha ao baixar imagem: {e})", file=sys.stderr)
    return blocos


def carregar_metricas_reais(xlsx_path: Path) -> dict[str, dict]:
    """Lê 'Desempenho dos seus anúncios.xlsx' → métricas reais por item_id (sem MLB)."""
    if not xlsx_path.exists():
        print(f"AVISO: relatório de métricas não encontrado em {xlsx_path}. "
              f"Seguindo sem métricas reais (gargalo ficará 'indefinido').", file=sys.stderr)
        return {}
    try:
        import openpyxl
    except ImportError:
        print("AVISO: openpyxl não instalado; seguindo sem métricas reais.", file=sys.stderr)
        return {}

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Relatório"] if "Relatório" in wb.sheetnames else wb.worksheets[0]
    linhas = list(ws.iter_rows(values_only=True))
    # acha a linha de cabeçalho (a que começa com "ID do anúncio")
    hdr_i = next((i for i, r in enumerate(linhas) if r and str(r[0]).strip() == "ID do anúncio"), None)
    if hdr_i is None:
        return {}
    def num(v):
        try:
            return float(str(v).replace("%", "").replace(".", "").replace(",", "."))
        except Exception:
            return 0.0

    metricas = {}
    for r in linhas[hdr_i + 1:]:
        if not r or not r[0] or str(r[0]).strip() in ("-", ""):
            continue
        item_id = str(r[0]).strip().replace("MLB", "")
        # colunas conforme o export: ID, Anúncio, Status, Variação, SKU, Visitas únicas,
        # Qtd vendas, Compradores únicos, Unidades vendidas, Vendas brutas, %part, Conv.vendas, Conv.compradores
        def g(i):
            return r[i] if i < len(r) else None
        linha = {
            "status": g(2),
            "visitas_unicas": g(5),
            "qtd_vendas": g(6),
            "compradores_unicos": g(7),
            "unidades_vendidas": g(8),
            "vendas_brutas": g(9),
            "conversao_visitas_vendas": g(11),
            "conversao_visitas_compradores": g(12),
        }
        # o relatório tem linhas duplicadas por anúncio (variações/subperíodos);
        # mantém a linha de MAIOR venda (é a que casa com a conversão do painel).
        antiga = metricas.get(item_id)
        if antiga is None or num(linha["qtd_vendas"]) > num(antiga["qtd_vendas"]):
            metricas[item_id] = linha
    wb.close()
    return metricas


def extrair_itens(dados) -> list[dict]:
    """Aceita os vários formatos que o n8n pode exportar:
    {items:[...]} | [{items:[...]}] | [{json:{items:[...]}}] | [item, item, ...]
    """
    if isinstance(dados, dict):
        if "items" in dados:
            return dados["items"]
        if "json" in dados and isinstance(dados["json"], dict):
            return dados["json"].get("items", [])
        return []
    if isinstance(dados, list):
        if dados and isinstance(dados[0], dict):
            primeiro = dados[0]
            if "items" in primeiro:
                return primeiro["items"]
            if "json" in primeiro and isinstance(primeiro["json"], dict) and "items" in primeiro["json"]:
                return primeiro["json"]["items"]
        return dados  # já é a lista de anúncios
    return []


def resumo_anuncio_txt(item: dict, metricas: dict | None) -> str:
    """Monta o bloco de texto com os dados do anúncio + métricas reais."""
    tipo = {"gold_pro": "Premium", "gold_special": "Clássico", "free": "Grátis"}.get(
        item.get("listing_type_id"), item.get("listing_type_id"))
    frete = (item.get("shipping") or {}).get("logistic_type")
    linhas = [
        f"item_id: {item['id']}",
        f"título atual: {item.get('title')}",
        f"preço: R$ {item.get('price')}  | tipo de anúncio: {tipo}  | frete/logística: {frete}",
        f"status: {item.get('status')}  | estoque: {item.get('available_quantity')}  | "
        f"catálogo: {item.get('catalog_listing')}",
        f"nº de fotos no anúncio: {len(item.get('pictures') or [])}  "
        f"(analisando as {min(MAX_IMGS_POR_ANUNCIO, len(item.get('pictures') or []))} primeiras)",
        "clip/vídeo: NÃO É POSSÍVEL SABER pela API (o campo video_id não reflete os Clips do ML). "
        "Vários anúncios desta conta JÁ TÊM clip. Nunca afirme que o anúncio não tem clip; se for "
        "recomendar, escreva a ação como 'criar clip SE ainda não houver' e deixe isso explícito.",
    ]
    if item.get("variations"):
        linhas.append(f"variações: {len(item['variations'])}")
    if metricas:
        linhas.append(
            "MÉTRICAS REAIS DO PAINEL ML: "
            f"visitas únicas={metricas.get('visitas_unicas')}, "
            f"vendas={metricas.get('qtd_vendas')}, "
            f"unidades={metricas.get('unidades_vendidas')}, "
            f"conversão (visitas→vendas)={metricas.get('conversao_visitas_vendas')}, "
            f"conversão (visitas→compradores)={metricas.get('conversao_visitas_compradores')}"
        )
    else:
        linhas.append("MÉTRICAS REAIS DO PAINEL ML: não disponível para este item.")

    linhas.append("\nFICHA TÉCNICA:")
    for a in (item.get("attributes") or []):
        if a.get("valor"):
            linhas.append(f"  - {a['nome']}: {a['valor']}")

    linhas.append("\nDESCRIÇÃO:")
    linhas.append((item.get("description") or "(vazia)")[:4000])

    perguntas = item.get("questions") or []
    if perguntas:
        linhas.append("\nPERGUNTAS DOS COMPRADORES (revelam o atrito real):")
        for q in perguntas[:15]:
            linhas.append(f"  P: {q.get('pergunta')}")
            if q.get("resposta"):
                linhas.append(f"  R: {q['resposta']}")
    else:
        linhas.append(
            "\nPERGUNTAS: não foi possível coletar (a API negou o acesso — erro 403 de permissão). "
            "Isso NÃO significa que o anúncio não tem perguntas. Não conclua nada sobre ausência de "
            "perguntas e deixe 'duvidas_recorrentes' vazio se não houver outra evidência."
        )
    return "\n".join(linhas)


# ----------------------------------------------------------------------------
# Chamadas ao Claude
# ----------------------------------------------------------------------------
def analisar_anuncio(client, system_blocks, item, metricas) -> dict:
    img_blocks = baixar_imagens(item.get("pictures") or [], MAX_IMGS_POR_ANUNCIO)
    texto = resumo_anuncio_txt(item, metricas)
    content = img_blocks + [{
        "type": "text",
        "text": (
            "Analise ESTE anúncio (as imagens acima são a galeria real dele) e responda no schema JSON.\n\n"
            + texto
        ),
    }]
    resp = client.messages.create(
        model=MODELO,
        max_tokens=MAX_TOKENS,
        system=system_blocks,
        thinking={"type": "adaptive"},
        output_config={"effort": EFFORT, "format": {"type": "json_schema", "schema": SCHEMA_ANUNCIO}},
        messages=[{"role": "user", "content": content}],
    )
    if resp.stop_reason == "refusal":
        raise RuntimeError(f"Recusa de segurança no item {item['id']}: {getattr(resp, 'stop_details', None)}")
    if resp.stop_reason == "max_tokens":
        raise RuntimeError(
            f"Resposta truncada no item {item['id']} (bateu MAX_TOKENS={MAX_TOKENS}). "
            f"Suba MAX_TOKENS ou baixe EFFORT."
        )
    texto_json = next((b.text for b in resp.content if b.type == "text"), "")
    return json.loads(texto_json)


def analisar_conta(client, system_blocks, resultados, metricas_por_id) -> dict:
    # resumo compacto por anúncio para caber no contexto
    linhas = []
    for r in resultados:
        m = metricas_por_id.get(r["item_id"].replace("MLB", ""), {})
        linhas.append(
            f"- {r['item_id']} | nota {r.get('nota_geral')} | gargalo {r.get('gargalo')} | "
            f"prioridade {r.get('prioridade')} | visitas={m.get('visitas_unicas')} "
            f"vendas={m.get('qtd_vendas')} conv={m.get('conversao_visitas_vendas')} | "
            f"{r.get('titulo_atual','')[:60]}"
        )
    prompt = (
        f"Abaixo está o resultado da análise individual de todos os anúncios de {CONTA}, "
        "com as métricas reais. Com base nisso e nas regras da Central, gere as JOGADAS DE CONTA "
        "(nível macro): cupons, promoções, kits, uso de Full, Mercado Ads, Minha Página/recompra, "
        "ajustes de preço, catálogo e reputação. Priorize e diga em quais itens aplicar.\n\n"
        + "\n".join(linhas)
    )
    resp = client.messages.create(
        model=MODELO,
        max_tokens=MAX_TOKENS,
        system=system_blocks,
        thinking={"type": "adaptive"},
        output_config={"effort": EFFORT, "format": {"type": "json_schema", "schema": SCHEMA_CONTA}},
        messages=[{"role": "user", "content": prompt}],
    )
    if resp.stop_reason == "refusal":
        raise RuntimeError("Recusa de segurança na análise de conta.")
    if resp.stop_reason == "max_tokens":
        raise RuntimeError(f"Plano de conta truncado (bateu MAX_TOKENS={MAX_TOKENS}). Suba MAX_TOKENS.")
    texto_json = next((b.text for b in resp.content if b.type == "text"), "")
    return json.loads(texto_json)


# ----------------------------------------------------------------------------
# Escrita dos relatórios
# ----------------------------------------------------------------------------
def escrever_plano_anuncios(resultados, caminho):
    resultados = sorted(resultados, key=lambda r: r.get("nota_geral", 0))
    L = [f"# Plano por anúncio — {CONTA}", ""]
    for r in resultados:
        np = r.get("notas_pilares", {})
        L.append(f"## {r['item_id']} — nota {r.get('nota_geral')} — prioridade {r.get('prioridade')}")
        L.append(f"*{r.get('titulo_atual','')}*")
        L.append("")
        L.append(f"- **Gargalo:** {r.get('gargalo')}")
        L.append(f"- **Notas:** imagens {np.get('imagens')} · título {np.get('titulo')} · "
                 f"ficha {np.get('ficha_tecnica')} · descrição {np.get('descricao')}")
        L.append(f"- **Diagnóstico:** {r.get('diagnostico')}")
        if r.get("problemas"):
            L.append("- **Problemas:**")
            for p in r["problemas"]:
                L.append(f"  - {p}")
        if r.get("duvidas_recorrentes"):
            L.append("- **Dúvidas recorrentes (levar p/ ficha/descrição):** " + "; ".join(r["duvidas_recorrentes"]))
        if r.get("titulo_sugerido"):
            L.append(f"- **Título sugerido:** `{r['titulo_sugerido']}`")
        L.append("- **Ações:**")
        for a in sorted(r.get("acoes", []), key=lambda x: x.get("ordem", 0)):
            L.append(f"  {a.get('ordem')}. [{a.get('tipo')}] {a.get('acao')} "
                     f"— _impacto:_ {a.get('impacto_esperado')} · _esforço:_ {a.get('esforco')}")
        L.append("")
    caminho.write_text("\n".join(L), encoding="utf-8")


def escrever_plano_conta(conta, caminho):
    L = [f"# Plano de conta — {CONTA}", "", conta.get("resumo", ""), ""]
    ordem = {"alta": 0, "media": 1, "baixa": 2}
    for j in sorted(conta.get("jogadas", []), key=lambda x: ordem.get(x.get("prioridade"), 3)):
        L.append(f"## [{j.get('tipo')}] {j.get('titulo')} — prioridade {j.get('prioridade')}")
        L.append(j.get("descricao", ""))
        if j.get("itens_alvo"):
            L.append(f"*Itens-alvo:* {', '.join(j['itens_alvo'])}")
        L.append(f"*Impacto esperado:* {j.get('impacto_esperado')}")
        L.append("")
    if conta.get("itens_prioritarios"):
        L.append("## Itens prioritários")
        L.append(", ".join(conta["itens_prioritarios"]))
    caminho.write_text("\n".join(L), encoding="utf-8")


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Analista de anúncios ML")
    ap.add_argument("--dados", type=Path, default=PADRAO_DADOS)
    ap.add_argument("--regras", type=Path, default=PADRAO_REGRAS)
    ap.add_argument("--metricas", type=Path, default=PADRAO_METRICAS)
    ap.add_argument("--limite", type=int, default=0, help="analisa só os N primeiros (teste)")
    ap.add_argument("--sem-conta", action="store_true", help="pula a análise de conta")
    args = ap.parse_args()

    if not args.dados.exists():
        sys.exit(f"ERRO: {args.dados} não existe. Rode o coletor no n8n e salve o JSON aí.")

    dados = json.loads(args.dados.read_text(encoding="utf-8"))
    itens = extrair_itens(dados)
    if not itens:
        sys.exit(f"ERRO: nenhum anúncio encontrado em {args.dados}. "
                 f"O JSON precisa ter a chave 'items' (saída do coletor n8n).")
    itens = [i for i in itens if i.get("id")]
    if args.limite:
        itens = itens[:args.limite]
    print(f"{len(itens)} anúncios carregados de {args.dados.name}")

    regras = args.regras.read_text(encoding="utf-8") if args.regras.exists() else "(regras não encontradas)"
    metricas_por_id = carregar_metricas_reais(args.metricas)

    client = anthropic.Anthropic()  # usa ANTHROPIC_API_KEY ou perfil `ant auth login`

    # System com cache (estável entre todas as chamadas → barato a partir da 2ª)
    system_blocks = [{
        "type": "text",
        "text": SYSTEM_PROMPT_BASE.format(regras=regras),
        "cache_control": {"type": "ephemeral"},
    }]

    SAIDA_DIR.mkdir(exist_ok=True)
    resultados = []
    for i, item in enumerate(itens, 1):
        mid = str(item["id"]).replace("MLB", "")
        print(f"[{i}/{len(itens)}] analisando {item['id']} — {item.get('title','')[:50]}...")
        try:
            r = analisar_anuncio(client, system_blocks, item, metricas_por_id.get(mid))
            resultados.append(r)
        except Exception as e:
            print(f"    ERRO no item {item['id']}: {e}", file=sys.stderr)
        time.sleep(0.5)  # respiro entre chamadas

    if not resultados:
        sys.exit("Nenhum anúncio analisado com sucesso.")

    (SAIDA_DIR / "resultados_brutos.json").write_text(
        json.dumps(resultados, ensure_ascii=False, indent=2), encoding="utf-8")
    escrever_plano_anuncios(resultados, SAIDA_DIR / "plano_por_anuncio.md")
    print(f"OK → {SAIDA_DIR / 'plano_por_anuncio.md'}")

    if not args.sem_conta:
        print("Gerando plano de conta (cupons/promos/kits/estrutura)...")
        try:
            conta = analisar_conta(client, system_blocks, resultados, metricas_por_id)
            (SAIDA_DIR / "plano_conta.json").write_text(
                json.dumps(conta, ensure_ascii=False, indent=2), encoding="utf-8")
            escrever_plano_conta(conta, SAIDA_DIR / "plano_conta.md")
            print(f"OK → {SAIDA_DIR / 'plano_conta.md'}")
        except Exception as e:
            print(f"    ERRO no plano de conta: {e}", file=sys.stderr)


if __name__ == "__main__":
    main()
